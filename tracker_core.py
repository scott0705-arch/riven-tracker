"""
tracker_core
============
Shared logic used by BOTH:
  - collect.py  (runs in the cloud, GitHub Actions, every hour)
  - client.py   (the desktop app you and your friends run)

Contains: warframe.market API access, weapon-name resolution, price
collection, statistics/analysis, and the Excel export. No GUI code and no
scheduling code lives here.
"""

import json
import os
import re
import sys
import tempfile
import time
import urllib.error
import urllib.parse
import urllib.request
from datetime import datetime, timezone

API_BASE = "https://api.warframe.market/v1"
USER_AGENT = "RivenPriceTracker/2.0 (personal price-tracking tool)"
REQUEST_GAP_SECONDS = 0.5          # stay well under warframe.market's ~3 req/s limit
HTTP_RETRIES = 3
ITEMS_CACHE_MAX_AGE_DAYS = 7

ANALYSIS_DEFAULTS = {
    "lookback": 168,               # recent samples used by the analysis (168 = 1 week)
    "sell_percentile": 0.5,        # where in the recent range you expect to sell
    "desired_profit": 50,          # platinum profit wanted per flip
    "safety_margin": 0.1,          # haircut on the projected sell price
}

ANALYSIS_HEADERS = ["Weapon", "Latest price", "Samples", "Recent min",
                    "Recent median", "Recent avg", "Projected sell",
                    "Good buy price", "Profit at good buy"]


# ---------------------------------------------------------------------------
# small utilities
# ---------------------------------------------------------------------------
def atomic_write_json(path, data):
    fd, tmp = tempfile.mkstemp(dir=os.path.dirname(os.path.abspath(path)),
                               suffix=".tmp")
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2)
        os.replace(tmp, path)
    except Exception:
        if os.path.exists(tmp):
            os.unlink(tmp)
        raise


def load_json(path, default=None):
    try:
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    except (FileNotFoundError, json.JSONDecodeError):
        return default


def parse_ts(ts):
    """Parse an ISO timestamp from history. Old entries are naive local time,
    new (server) entries are UTC with an offset - handle both, return an
    aware-or-naive datetime suitable for sorting/formatting."""
    dt = datetime.fromisoformat(ts)
    return dt


def ts_local_str(ts):
    """Human-readable local time for a history timestamp."""
    dt = parse_ts(ts)
    if dt.tzinfo is not None:
        dt = dt.astimezone()           # convert UTC (server) -> viewer's local
    return dt.strftime("%Y-%m-%d %H:%M")


def utcnow_iso():
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


# ---------------------------------------------------------------------------
# history helpers (operate on a plain snapshot list)
# ---------------------------------------------------------------------------
def load_history_file(path):
    data = load_json(path, {}) or {}
    snaps = data.get("snapshots", [])
    return sorted(snaps, key=lambda s: s["timestamp"])


def append_history_file(path, snapshot):
    snaps = load_history_file(path)
    if any(s["timestamp"] == snapshot["timestamp"] for s in snaps):
        return snaps
    snaps.append(snapshot)
    snaps.sort(key=lambda s: s["timestamp"])
    atomic_write_json(path, {"snapshots": snaps})
    return snaps

RETENTION_DAYS = 14

def prune_history_file(path, days=RETENTION_DAYS):
    """Drop snapshots older than `days`. Returns the kept list."""
    from datetime import timedelta
    snaps = load_history_file(path)
    cutoff = datetime.now(timezone.utc) - timedelta(days=days)
    def keep(s):
        dt = parse_ts(s["timestamp"])
        if dt.tzinfo is None:                     # old local-time entries
            dt = dt.replace(tzinfo=timezone.utc)  # treat as UTC; close enough
        return dt >= cutoff
    kept = [s for s in snaps if keep(s)]
    if len(kept) != len(snaps):
        atomic_write_json(path, {"snapshots": kept})
    return kept

def all_weapon_names(weapons, history):
    """Config weapons first (in order), then any extra names found in history
    (e.g. weapons that were removed later) so their data stays visible."""
    names = [w["name"] for w in weapons]
    seen = set(names)
    for s in history:
        for n in s["prices"]:
            if n not in seen:
                names.append(n)
                seen.add(n)
    return names


# ---------------------------------------------------------------------------
# HTTP
# ---------------------------------------------------------------------------
def fetch_json(url, platform="pc", log=None):
    """GET a warframe.market API url with retries. Returns parsed JSON."""
    last_err = None
    for attempt in range(HTTP_RETRIES):
        try:
            req = urllib.request.Request(url, headers={
                "User-Agent": USER_AGENT,
                "Accept": "application/json",
                "Platform": platform,
                "Language": "en",
            })
            with urllib.request.urlopen(req, timeout=30) as resp:
                return json.loads(resp.read().decode("utf-8"))
        except urllib.error.HTTPError as e:
            last_err = e
            if e.code == 429:                      # rate limited: back off harder
                time.sleep(5 * (attempt + 1))
            elif 500 <= e.code < 600:
                time.sleep(3 * (attempt + 1))
            else:
                raise                              # other 4xx: no point retrying
        except (urllib.error.URLError, TimeoutError, json.JSONDecodeError) as e:
            last_err = e
            time.sleep(3 * (attempt + 1))
    raise RuntimeError(f"request failed after {HTTP_RETRIES} attempts: {url} ({last_err})")


# ---------------------------------------------------------------------------
# riven item name resolution
# ---------------------------------------------------------------------------
def slugify(name):
    s = name.strip().lower()
    s = s.replace("&", "and").replace("'", "").replace("-", "_")
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^a-z0-9_]", "", s)
    return s


def get_riven_items(cache_file, force_refresh=False, log=None):
    """Return list of {"item_name","url_name"} for all riven-capable weapons.
    Cached locally; refreshed every ITEMS_CACHE_MAX_AGE_DAYS."""
    if not force_refresh:
        cache = load_json(cache_file, {}) or {}
        age = time.time() - cache.get("fetched_at", 0)
        if age < ITEMS_CACHE_MAX_AGE_DAYS * 86400 and cache.get("items"):
            return cache["items"]
    try:
        items = None
        try:  # current endpoint (v2)
            data = fetch_json("https://api.warframe.market/v2/riven/weapons")
            items = [{"item_name": w["i18n"]["en"]["name"], "url_name": w["slug"]}
                     for w in data["data"]]
        except Exception:
            data = fetch_json(f"{API_BASE}/riven/items")
            items = [{"item_name": i["item_name"], "url_name": i["url_name"]}
                     for i in data["payload"]["items"]]
        if not items:
            raise RuntimeError("empty weapon list")
        atomic_write_json(cache_file, {"fetched_at": time.time(), "items": items})
        return items
    except Exception:
        cache = load_json(cache_file, {}) or {}
        return cache.get("items", [])


def resolve_weapon(name, items):
    """Map a user-typed weapon name to (display_name, url_name).
    Returns (result_or_None, suggestions)."""
    target = slugify(name)
    for it in items:
        if it["url_name"] == target or slugify(it["item_name"]) == target:
            return (it["item_name"], it["url_name"]), []
    import difflib
    matches = difflib.get_close_matches(
        name.strip().title(), [i["item_name"] for i in items], n=3, cutoff=0.6)
    if not items:                     # offline fallback: trust the user's spelling
        return (name.strip().title(), target), []
    return None, matches


# ---------------------------------------------------------------------------
# price collection
# ---------------------------------------------------------------------------
def lowest_ingame_price(url_name, platform="pc", max_rerolls=0):
    """Lowest buyout price among visible, open, direct-sale riven auctions with
    at most `max_rerolls` rerolls, whose seller is currently in-game.
    Returns int or None if no such listing."""
    qs = urllib.parse.urlencode({
        "type": "riven",
        "weapon_url_name": url_name,
        "buyout_policy": "direct",     # direct sales only -- auctions ignored
        "re_rolls_min": 0,
        "re_rolls_max": max_rerolls,   # server-side reroll filter
        "sort_by": "price_asc",
    })
    data = fetch_json(f"{API_BASE}/auctions/search?{qs}", platform=platform)
    prices = []
    for a in data.get("payload", {}).get("auctions", []):
        if a.get("closed") or not a.get("visible", True):
            continue
        if a.get("buyout_price") is None:
            continue
        item = a.get("item") or {}
        if item.get("re_rolls", 0) > max_rerolls:   # belt-and-braces re-check
            continue
        owner = a.get("owner") or {}
        if owner.get("status") != "ingame":
            continue
        prices.append(a["buyout_price"])
    return min(prices) if prices else None


def collect_snapshot(weapons, platform="pc", max_rerolls=0, progress=None):
    """Fetch current prices for every weapon in `weapons`.
    Returns {"timestamp": utc_iso, "prices": {display_name: int|None}}."""
    prices = {}
    for i, w in enumerate(weapons):
        try:
            p = lowest_ingame_price(w["url_name"], platform, max_rerolls)
        except Exception as e:
            if progress:
                progress(f"{w['name']}: fetch failed ({e})")
            p = None
        prices[w["name"]] = p
        if progress:
            progress(f"{w['name']}: {p if p is not None else 'no in-game listing'}"
                     f"  ({i + 1}/{len(weapons)})")
        if i < len(weapons) - 1:
            time.sleep(REQUEST_GAP_SECONDS)
    return {"timestamp": utcnow_iso(), "prices": prices}


# ---------------------------------------------------------------------------
# analysis
# ---------------------------------------------------------------------------
def percentile_inc(values, p):
    """Excel PERCENTILE.INC (linear interpolation) on a list of numbers."""
    v = sorted(values)
    if not v:
        return None
    if len(v) == 1:
        return float(v[0])
    rank = p * (len(v) - 1)
    lo = int(rank)
    frac = rank - lo
    if lo + 1 >= len(v):
        return float(v[-1])
    return v[lo] + (v[lo + 1] - v[lo]) * frac


def compute_analysis(weapons, history, analysis):
    """Per-weapon stats from history. Returns list of tuples matching
    ANALYSIS_HEADERS. `analysis` = dict with lookback / sell_percentile /
    desired_profit / safety_margin (the CLIENT-side tunables)."""
    a = {**ANALYSIS_DEFAULTS, **(analysis or {})}
    lookback = int(a["lookback"])
    pctl = float(a["sell_percentile"])
    profit_target = float(a["desired_profit"])
    margin = float(a["safety_margin"])

    rows = []
    for name in all_weapon_names(weapons, history):
        series = [s["prices"].get(name) for s in history if name in s["prices"]]
        numeric = [v for v in series if isinstance(v, (int, float))]
        window = [v for v in series[-lookback:] if isinstance(v, (int, float))]
        latest = series[-1] if series else None
        latest_disp = latest if isinstance(latest, (int, float)) \
            else ("no listing" if series else "")
        if window:
            sell = percentile_inc(window, pctl) * (1 - margin)
            good_buy = max(0, round(sell - profit_target))
            rows.append((name, latest_disp, len(numeric), min(window),
                         percentile_inc(window, 0.5),
                         round(sum(window) / len(window), 1),
                         round(sell), good_buy, round(sell) - good_buy))
        else:
            rows.append((name, latest_disp, len(numeric), "", "", "", "", "", ""))
    return rows


# ---------------------------------------------------------------------------
# Excel export (client-side convenience; regenerated in full each time)
# ---------------------------------------------------------------------------
DATA_SHEET = "Data"
ANALYSIS_SHEET = "Analysis"
SETTINGS_SHEET = "Settings"
TS_FORMAT = "yyyy-mm-dd hh:mm"

SETTINGS_ROWS = [
    ("Lookback window (number of recent samples)", "lookback",
     "How many of the most recent hourly samples the analysis uses. 168 = 1 week."),
    ("Projected sell percentile (0 - 1)", "sell_percentile",
     "Where in the recent price range you expect to sell. 0.5 = median of recent lows."),
    ("Desired profit (platinum)", "desired_profit",
     "Profit you want per flip. Subtracted from projected sell price."),
    ("Safety margin (0 - 1)", "safety_margin",
     "Extra haircut on the projected sell price. 0.1 = assume you sell 10% below it."),
]


def export_workbook(path, weapons, history, analysis):
    """Rebuild the entire workbook from history. Returns True on success,
    False if the file is locked (open in Excel)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    st = {
        "header": Font(name="Arial", bold=True),
        "base": Font(name="Arial"),
        "input": Font(name="Arial", color="0000FF"),
        "grey": PatternFill("solid", fgColor="DDDDDD"),
    }
    wb = Workbook()

    # --- Data sheet: rows = weapons, columns = snapshots -------------------
    ws = wb.active
    ws.title = DATA_SHEET
    ws["A1"] = "Weapon"
    ws["A1"].font = st["header"]
    ws["A1"].fill = st["grey"]
    ws.column_dimensions["A"].width = 28
    ws.freeze_panes = "B2"

    names = all_weapon_names(weapons, history)
    for i, name in enumerate(names):
        ws.cell(row=i + 2, column=1, value=name).font = st["base"]
    for c, snap in enumerate(history, start=2):
        dt = parse_ts(snap["timestamp"])
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)   # Excel wants naive local
        head = ws.cell(row=1, column=c, value=dt)
        head.number_format = TS_FORMAT
        head.font = st["header"]
        head.fill = st["grey"]
        ws.column_dimensions[get_column_letter(c)].width = 16
        for i, name in enumerate(names):
            v = snap["prices"].get(name)
            if isinstance(v, (int, float)):
                ws.cell(row=i + 2, column=c, value=v).font = st["base"]

    # --- Analysis sheet ----------------------------------------------------
    aws = wb.create_sheet(ANALYSIS_SHEET)
    for c, h in enumerate(ANALYSIS_HEADERS, start=1):
        cell = aws.cell(row=1, column=c, value=h)
        cell.font = st["header"]
        cell.fill = st["grey"]
    aws.freeze_panes = "A2"
    aws.column_dimensions["A"].width = 28
    for col in "BCDEFGHI":
        aws.column_dimensions[col].width = 15
    for r, row in enumerate(compute_analysis(weapons, history, analysis), start=2):
        for c, v in enumerate(row, start=1):
            aws.cell(row=r, column=c, value=v).font = st["base"]

    # --- Settings sheet ----------------------------------------------------
    s = wb.create_sheet(SETTINGS_SHEET)
    s["A1"] = "Buy-price settings  -  managed in the tracker app; change them there"
    s["A1"].font = st["header"]
    a = {**ANALYSIS_DEFAULTS, **(analysis or {})}
    for i, (label, key, note) in enumerate(SETTINGS_ROWS):
        r = 2 + i
        s.cell(row=r, column=1, value=label).font = st["base"]
        c = s.cell(row=r, column=2, value=a[key])
        c.font = st["input"]
        s.cell(row=r, column=3, value=note).font = st["base"]
    s["A7"] = "Good buy price = projected sell price x (1 - safety margin) - desired profit"
    s["A7"].font = st["base"]
    s.column_dimensions["A"].width = 44
    s.column_dimensions["B"].width = 10
    s.column_dimensions["C"].width = 70

    # --- atomic save; report locked file ------------------------------------
    out_dir = os.path.dirname(os.path.abspath(path))
    fd, tmp = tempfile.mkstemp(dir=out_dir, suffix=".xlsx")
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, path)
        return True
    except PermissionError:
        return False
    finally:
        if os.path.exists(tmp):
            try:
                os.unlink(tmp)
            except OSError:
                pass
